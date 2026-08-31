"""
backend/tools/xlsx_report.py
-----------------------------
Industrial XLSX / Spreadsheet report generation tool with formatting and styling.

Security & Integrity:
- Writes strictly within data/sandbox/
- Atomic file generation via temporary files
- Rejection of path traversal and unsafe filenames
- Computes SHA-256 artifact hash for cryptographic audit logging
- Full support for openpyxl styling (headers, borders, color-coded compliance status)
"""

from __future__ import annotations

import hashlib
import logging
from pathlib import Path
from typing import Any, Dict, List, Optional

from pydantic import BaseModel, Field

from backend.tools.safety import atomic_write_file, sanitize_filename, validate_path_within

logger = logging.getLogger(__name__)


class XlsxReportInput(BaseModel):
    """Input schema for generating structured Excel audit & compliance reports."""
    filename: str = Field(
        ...,
        min_length=1,
        max_length=255,
        description="Target spreadsheet filename ending in .xlsx, e.g. 'mrpl_compliance_report.xlsx'.",
    )
    title: str = Field(
        ...,
        min_length=3,
        description="Report title, e.g. 'MRPL Refinery Hydrocarbon Stream Lab Compliance Report'.",
    )
    headers: List[str] = Field(
        ...,
        min_length=1,
        description="List of column headers, e.g. ['Sample ID', 'Parameter', 'Measured Value', 'Standard', 'Status'].",
    )
    rows: List[List[Any]] = Field(
        ...,
        description="Matrix of data rows matching the header columns.",
    )
    summary_notes: Optional[str] = Field(
        default="",
        description="Executive summary notes or regulatory citations included in the header block.",
    )
    overwrite: bool = Field(
        default=True,
        description="Whether to overwrite if file exists.",
    )


def create_xlsx_report(sandbox_dir: Path) -> callable:
    """Create the execute_xlsx_report function bound to sandbox_dir."""

    async def execute_xlsx_report(args: XlsxReportInput) -> Dict[str, Any]:
        # Validate and sanitize filename
        safe_name = sanitize_filename(args.filename)
        if not safe_name.lower().endswith(".xlsx"):
            safe_name += ".xlsx"

        target_path = validate_path_within(safe_name, sandbox_dir)

        # Build workbook using openpyxl
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
            from openpyxl.utils import get_column_letter

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Compliance Report"
            ws.views.sheetView[0].showGridLines = True

            # Color palette (Enterprise Industrial Theme)
            header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")  # Slate-800
            header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            title_font = Font(name="Calibri", size=14, bold=True, color="0F172A")
            meta_font = Font(name="Calibri", size=10, italic=True, color="475569")
            regular_font = Font(name="Calibri", size=10, color="1E293B")
            pass_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")   # Emerald-100
            pass_font = Font(name="Calibri", size=10, bold=True, color="166534")
            fail_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")   # Rose-100
            fail_font = Font(name="Calibri", size=10, bold=True, color="991B1B")

            thin_border = Border(
                left=Side(style="thin", color="CBD5E1"),
                right=Side(style="thin", color="CBD5E1"),
                top=Side(style="thin", color="CBD5E1"),
                bottom=Side(style="thin", color="CBD5E1"),
            )

            # Row 1: Title
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(args.headers), 4))
            ws.cell(row=1, column=1, value=args.title).font = title_font

            # Row 2: Metadata / Notes
            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(len(args.headers), 4))
            ws.cell(
                row=2, column=1,
                value=f"Generated via Sovereign Agentic Workbench | {args.summary_notes or 'Automated Diligence & Verification'}"
            ).font = meta_font

            # Row 4: Headers
            header_row = 4
            for col_idx, header in enumerate(args.headers, 1):
                cell = ws.cell(row=header_row, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border

            # Data rows
            current_row = header_row + 1
            for r_data in args.rows:
                for col_idx, val in enumerate(r_data, 1):
                    cell = ws.cell(row=current_row, column=col_idx, value=val)
                    cell.font = regular_font
                    cell.border = thin_border

                    # Highlight compliance columns
                    val_str = str(val).upper()
                    if val_str in ("COMPLIANT", "PASS", "SAFE", "NORMAL"):
                        cell.fill = pass_fill
                        cell.font = pass_font
                    elif val_str in ("DEVIATION", "FAIL", "ALERT", "CRITICAL", "EXCEEDED"):
                        cell.fill = fail_fill
                        cell.font = fail_font

                current_row += 1

            # Auto-adjust column widths
            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    val_str = str(cell.value or "")
                    if len(val_str) > max_len and cell.row > 2:
                        max_len = len(val_str)
                ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

            # Write out to temp file and rename atomically
            temp_target = target_path.with_suffix(".tmp.xlsx")
            wb.save(temp_target)
            wb.close()

            # Atomic replace
            import os
            os.replace(temp_target, target_path)

        except Exception as exc:
            logger.error("XlsxReport generation failed: %s", exc)
            raise RuntimeError(f"Failed to create XLSX report '{safe_name}': {exc}")

        # Compute SHA-256 hash of generated artifact
        file_bytes = target_path.read_bytes()
        sha256_hash = hashlib.sha256(file_bytes).hexdigest()
        rel_path = f"data/sandbox/{safe_name}"

        logger.info(
            "xlsx_report_created | path=%s rows=%d size=%d hash=%s",
            rel_path, len(args.rows), len(file_bytes), sha256_hash[:16]
        )

        return {
            "created_path": rel_path,
            "filename": safe_name,
            "size_bytes": len(file_bytes),
            "row_count": len(args.rows),
            "column_count": len(args.headers),
            "sha256_hash": sha256_hash,
            "format": "xlsx",
        }

    return execute_xlsx_report
